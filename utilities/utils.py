import inspect
import softest
import logging
from openpyxl import Workbook, load_workbook
import csv



class Utils(softest.TestCase):
    def assertListItemText(self,list,value):

        for item in list:
            print("the text is: "+ item.text)
            self.soft_assert(self.assertEqual, item.text,value)
            # assert item.text == value
            if item.text == value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()

    def custom_logger( loglevel = logging.DEBUG):

        logger_name = inspect.stack()[1][3]
        
        # create logger
        logger = logging.getLogger(logger_name)#as apposed to using a stream __name__ is a best practice
        logger.setLevel(loglevel)


        # create console  or FileHandler handler and set level to debug
 

        fh = logging.FileHandler("automation.log", mode = "a")
        # fh.setLevel(logging.DEBUG)

        # create formatter
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s', datefmt= '%a %m/%d/%Y %I:%M:%S %p')

        # add formatter to ch
        fh.setFormatter(formatter)

        # add ch to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet = "Sheet1"):
        datalist = []
        wb = load_workbook(filename = file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column


        for i in range(2, 1 + row_ct):
            row = []
            for j in range(1, 1 + col_ct):
                row.append(sh.cell(row = i, column = j).value)
            datalist.append(row)

        return datalist

    def read_dat_from_csv(file_name):
        #create empty list
        datalist = []
        #open csvfile
        csvdata = open(file_name,"r")

        #create CSV reader
        reader = csv.reader(csvdata)
        #skip the header(first row)
        next(reader)
        #add the csv rows to list
        for rows in reader:
            datalist.append(rows)
        #return filled list
        return datalist

            

